#!/usr/bin/env python3
"""
生成测试用例Excel表格
包含所有测试用例的Input、Output和Expected Result
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_test_cases_excel():
    """创建测试用例Excel表格"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # 定义样式
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置表头
    headers = [
        'ID',
        'Service',
        'Test Case Name',
        'API Endpoint',
        'Method',
        'Input Parameters',
        'Expected Output',
        'Expected Status Code',
        'Skip',
        'Skip Reason',
        'Priority'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 定义测试用例数据
    test_cases = [
        # 用户管理服务测试用例
        {
            'id': 'UM-001',
            'service': 'User Management',
            'name': 'Health Check',
            'endpoint': '/health',
            'method': 'GET',
            'input': 'None',
            'expected_output': '{"status": "ok"} or "OK"',
            'status_code': '200',
            'skip': 'No',
            'skip_reason': '',
            'priority': 'Critical'
        },
        {
            'id': 'UM-002',
            'service': 'User Management',
            'name': 'Send Verification Code',
            'endpoint': '/auth/send-verification-code',
            'method': 'POST',
            'input': '{\n  "email": "test@example.com",\n  "recaptcha_token": "token"\n}',
            'expected_output': '{\n  "success": true,\n  "message": "Verification code sent"\n}',
            'status_code': '200, 400, 403',
            'skip': 'Yes',
            'skip_reason': '需要真实的邮箱',
            'priority': 'Critical'
        },
        {
            'id': 'UM-003',
            'service': 'User Management',
            'name': 'Verify Code',
            'endpoint': '/auth/verify-code',
            'method': 'POST',
            'input': '{\n  "email": "test@example.com",\n  "code": "123456",\n  "recaptcha_token": "token"\n}',
            'expected_output': '{\n  "success": true,\n  "token": "auth_token"\n}',
            'status_code': '200, 400, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的验证码',
            'priority': 'Critical'
        },
        {
            'id': 'UM-004',
            'service': 'User Management',
            'name': 'Admin Login',
            'endpoint': '/api/admin-login',
            'method': 'POST',
            'input': '{\n  "email": "admin@test.com",\n  "password": "password"\n}',
            'expected_output': '{\n  "success": true,\n  "token": "admin_token"\n}',
            'status_code': '200, 401, 403',
            'skip': 'Yes',
            'skip_reason': '需要真实的管理员钱包地址和签名',
            'priority': 'Normal'
        },
        {
            'id': 'UM-005',
            'service': 'User Management',
            'name': 'Verify Token',
            'endpoint': '/verify_token',
            'method': 'POST',
            'input': '{\n  "token": "test_token",\n  "email": "test@example.com"\n}',
            'expected_output': '{\n  "success": true,\n  "valid": true\n}',
            'status_code': '200, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的token',
            'priority': 'Normal'
        },
        {
            'id': 'UM-006',
            'service': 'User Management',
            'name': 'Get Users by Wallet',
            'endpoint': '/api/get-users',
            'method': 'GET',
            'input': 'Params:\n{\n  "wallet_address": "wallet123",\n  "chain": "solana",\n  "recaptcha_token": "token"\n}',
            'expected_output': '{\n  "success": true,\n  "user": {...}\n}',
            'status_code': '200, 400, 401, 403',
            'skip': 'No',
            'skip_reason': '',
            'priority': 'Critical'
        },
        {
            'id': 'UM-007',
            'service': 'User Management',
            'name': 'Get Users by Email',
            'endpoint': '/api/get-users',
            'method': 'GET',
            'input': 'Params:\n{\n  "email": "test@example.com",\n  "recaptcha_token": "token"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "success": true,\n  "user": {...}\n}',
            'status_code': '200, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的token',
            'priority': 'Critical'
        },
        {
            'id': 'UM-008',
            'service': 'User Management',
            'name': 'Update User Info',
            'endpoint': '/api/update-user-info',
            'method': 'POST',
            'input': '{\n  "email": "test@example.com",\n  "user_name": "Test User",\n  "telegram_id": "@testuser",\n  "twitter_id": "@testuser"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "success": true,\n  "message": "User info updated"\n}',
            'status_code': '200, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的token',
            'priority': 'Normal'
        },
        {
            'id': 'UM-009',
            'service': 'User Management',
            'name': 'Disconnect Wallet',
            'endpoint': '/api/disconnect-wallet-from-email',
            'method': 'POST',
            'input': '{\n  "email": "test@example.com",\n  "chain": "solana"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "success": true,\n  "message": "Wallet disconnected"\n}',
            'status_code': '200, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的token',
            'priority': 'Normal'
        },
        {
            'id': 'UM-010',
            'service': 'User Management',
            'name': 'Daily Login Points',
            'endpoint': '/api/daily-login-points',
            'method': 'POST',
            'input': '{\n  "email": "test@example.com"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "success": true,\n  "points": 10\n}',
            'status_code': '200, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的token',
            'priority': 'Normal'
        },
        {
            'id': 'UM-011',
            'service': 'User Management',
            'name': 'Get Point Transactions',
            'endpoint': '/api/get-point-transactions',
            'method': 'GET',
            'input': 'Params:\n{\n  "email": "test@example.com",\n  "token": "test_token"\n}',
            'expected_output': '{\n  "success": true,\n  "transactions": [...]\n}',
            'status_code': '200, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的token',
            'priority': 'Normal'
        },
        {
            'id': 'UM-012',
            'service': 'User Management',
            'name': 'Create Stripe Checkout',
            'endpoint': '/api/create-stripe-checkout-session',
            'method': 'POST',
            'input': '{\n  "email": "test@example.com",\n  "price_id": "price_123",\n  "quantity": 1\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "success": true,\n  "session_id": "session_123"\n}',
            'status_code': '200, 400, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的token和数据',
            'priority': 'Normal'
        },
        {
            'id': 'UM-013',
            'service': 'User Management',
            'name': 'Record Crypto Payment',
            'endpoint': '/api/record-crypto-payment',
            'method': 'POST',
            'input': '{\n  "email": "test@example.com",\n  "wallet_address": "wallet123",\n  "tx_signature": "tx_sig",\n  "chain": "solana",\n  "tier": "amplifier"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "success": true,\n  "message": "Payment recorded"\n}',
            'status_code': '200, 400, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的交易数据',
            'priority': 'Normal'
        },
        {
            'id': 'UM-014',
            'service': 'User Management',
            'name': 'Reserve Robot',
            'endpoint': '/api/reserve-robot',
            'method': 'POST',
            'input': '{\n  "email": "test@example.com",\n  "robot_id": "arm1",\n  "reserve_date": "2025-10-10",\n  "reserve_time": "14:00"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "success": true,\n  "reservation_id": "res_123"\n}',
            'status_code': '200, 400, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的token和数据',
            'priority': 'Normal'
        },
        {
            'id': 'UM-015',
            'service': 'User Management',
            'name': 'Get Reserve Robot Info',
            'endpoint': '/api/get-reserve-robot-info',
            'method': 'GET',
            'input': 'Params:\n{\n  "email": "test@example.com"\n}',
            'expected_output': '{\n  "reservations": [...]\n}',
            'status_code': '200, 400',
            'skip': 'Yes',
            'skip_reason': '需要真实的邮箱',
            'priority': 'Normal'
        },
        {
            'id': 'UM-016',
            'service': 'User Management',
            'name': 'Get Admin Whitelist',
            'endpoint': '/api/get-admin-whitelist',
            'method': 'GET',
            'input': 'None',
            'expected_output': '{\n  "success": true,\n  "data": [\n    {\n      "admin_id": 123,\n      "wallet_address": "wallet123"\n    }\n  ]\n}',
            'status_code': '200, 401',
            'skip': 'No',
            'skip_reason': '',
            'priority': 'Normal'
        },
        {
            'id': 'UM-017',
            'service': 'User Management',
            'name': 'Set Live Paused',
            'endpoint': '/api/set-live-paused',
            'method': 'POST',
            'input': '{\n  "robot_id": "arm1",\n  "live_paused": true\n}\nHeaders:\n{\n  "Authorization": "Bearer admin_token"\n}',
            'expected_output': '{\n  "success": true,\n  "message": "Status updated"\n}',
            'status_code': '200, 401, 403',
            'skip': 'Yes',
            'skip_reason': '需要管理员权限',
            'priority': 'Normal'
        },
        {
            'id': 'UM-018',
            'service': 'User Management',
            'name': 'Get Blockchain Config',
            'endpoint': '/get_blockchain_config_address',
            'method': 'GET',
            'input': 'None',
            'expected_output': '{\n  "prismax_receiving_address": "addr123",\n  "verified_token_solana_usdc": "token123",\n  "live_paused_status": [...]\n}',
            'status_code': '200',
            'skip': 'No',
            'skip_reason': '',
            'priority': 'Normal'
        },
        {
            'id': 'UM-019',
            'service': 'User Management',
            'name': 'Get User Stats',
            'endpoint': '/api/user-stats',
            'method': 'GET',
            'input': 'Params:\n{\n  "email": "test@example.com"\n}',
            'expected_output': '{\n  "stats": {...}\n}',
            'status_code': '200, 400',
            'skip': 'Yes',
            'skip_reason': '需要真实的邮箱',
            'priority': 'Normal'
        },
        {
            'id': 'UM-020',
            'service': 'User Management',
            'name': 'Quiz Check Status',
            'endpoint': '/api/quiz/check-status',
            'method': 'GET',
            'input': 'Params:\n{\n  "email": "test@example.com"\n}',
            'expected_output': '{\n  "completed": true/false\n}',
            'status_code': '200, 400',
            'skip': 'Yes',
            'skip_reason': '需要真实的邮箱',
            'priority': 'Minor'
        },
        
        # Tele-Op服务测试用例
        {
            'id': 'TO-001',
            'service': 'Tele-Op',
            'name': 'Get Robots Status',
            'endpoint': '/robots/status',
            'method': 'GET',
            'input': 'None',
            'expected_output': '{\n  "robots": [\n    {\n      "robot_id": "arm1",\n      "is_available": true,\n      "queue_length": 5,\n      "youtube_stream_id": "stream123",\n      "live_paused": false\n    }\n  ]\n}',
            'status_code': '200',
            'skip': 'No',
            'skip_reason': '',
            'priority': 'Critical'
        },
        {
            'id': 'TO-002',
            'service': 'Tele-Op',
            'name': 'Get Live Paused Status',
            'endpoint': '/get_live_paused_status',
            'method': 'GET',
            'input': 'None',
            'expected_output': '{\n  "robots": [\n    {\n      "robot_id": "arm1",\n      "live_paused": false\n    }\n  ]\n}',
            'status_code': '200',
            'skip': 'No',
            'skip_reason': '',
            'priority': 'Normal'
        },
        {
            'id': 'TO-003',
            'service': 'Tele-Op',
            'name': 'Queue Status',
            'endpoint': '/queue/status',
            'method': 'POST',
            'input': '{\n  "userId": "user123",\n  "robotId": "arm1"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "queue": [...],\n  "robotAvailable": true,\n  "yourTurn": false\n}',
            'status_code': '200, 400, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的用户ID和token',
            'priority': 'Critical'
        },
        {
            'id': 'TO-004',
            'service': 'Tele-Op',
            'name': 'Join Queue',
            'endpoint': '/queue/join',
            'method': 'POST',
            'input': '{\n  "userId": "user123",\n  "robotId": "arm1"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "position": 3,\n  "message": "Joined queue"\n}',
            'status_code': '200, 400, 401, 403, 404',
            'skip': 'Yes',
            'skip_reason': '需要真实的用户ID和token',
            'priority': 'Critical'
        },
        {
            'id': 'TO-005',
            'service': 'Tele-Op',
            'name': 'Leave Queue',
            'endpoint': '/queue/leave',
            'method': 'POST',
            'input': '{\n  "userId": "user123",\n  "robotId": "arm1",\n  "inactive": false\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "message": "Left queue"\n}',
            'status_code': '200, 400, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的用户ID和token',
            'priority': 'Normal'
        },
        {
            'id': 'TO-006',
            'service': 'Tele-Op',
            'name': 'Use Robot',
            'endpoint': '/use_robot',
            'method': 'POST',
            'input': '{\n  "userId": "user123",\n  "robotId": "arm1"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "encrypted_url": "url",\n  "control_token": "token",\n  "expires": 300,\n  "isFirstTime": false\n}',
            'status_code': '200, 401, 403, 404',
            'skip': 'Yes',
            'skip_reason': '需要真实的用户ID和token，且用户需要在活跃队列中',
            'priority': 'Critical'
        },
        {
            'id': 'TO-007',
            'service': 'Tele-Op',
            'name': 'Get Session URL Decrypt Key',
            'endpoint': '/session_url_decrypt_key',
            'method': 'POST',
            'input': '{\n  "userId": "user123"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "key_b64": "base64_key",\n  "exp": 1234567890\n}',
            'status_code': '200, 401, 403, 500',
            'skip': 'Yes',
            'skip_reason': '需要真实的用户ID和token',
            'priority': 'Normal'
        },
        {
            'id': 'TO-008',
            'service': 'Tele-Op',
            'name': 'Get Control History',
            'endpoint': '/user/control_history',
            'method': 'POST',
            'input': '{\n  "userId": "user123"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "sessions": [...],\n  "total_controlled_hours": 10.5,\n  "total_reward_points": 1000\n}',
            'status_code': '200, 400, 401',
            'skip': 'Yes',
            'skip_reason': '需要真实的用户ID和token',
            'priority': 'Normal'
        },
        {
            'id': 'TO-009',
            'service': 'Tele-Op',
            'name': 'Fetch Session Complete Status',
            'endpoint': '/fetch_tele_op_session_complete_status',
            'method': 'POST',
            'input': '{\n  "userId": "user123",\n  "controlToken": "control_token"\n}\nHeaders:\n{\n  "Authorization": "Bearer token"\n}',
            'expected_output': '{\n  "status": "completed",\n  "result": {...}\n}\nor\n{\n  "status": "not_finished"\n}',
            'status_code': '200, 400, 401, 404',
            'skip': 'Yes',
            'skip_reason': '需要真实的用户ID、token和控制token',
            'priority': 'Normal'
        },
        {
            'id': 'TO-010',
            'service': 'Tele-Op',
            'name': 'Get Leaderboard',
            'endpoint': '/tele_op/leaderboard',
            'method': 'GET',
            'input': 'None',
            'expected_output': '{\n  "leaderboard": [\n    {\n      "user_id": "123",\n      "total_points": 5000,\n      "rank": 1,\n      "controlled_hours": 50.5\n    }\n  ]\n}',
            'status_code': '200, 500',
            'skip': 'No',
            'skip_reason': '',
            'priority': 'Normal'
        },
        {
            'id': 'TO-011',
            'service': 'Tele-Op',
            'name': 'Update Leaderboard',
            'endpoint': '/tele_op/leaderboard/update',
            'method': 'POST',
            'input': 'Headers:\n{\n  "Authorization": "Bearer internal_token"\n}',
            'expected_output': '{\n  "status": "ok"\n}',
            'status_code': '200, 401, 500',
            'skip': 'Yes',
            'skip_reason': '需要真实的内部API token',
            'priority': 'Minor'
        },
        {
            'id': 'TO-012',
            'service': 'Tele-Op',
            'name': 'Robot Disconnect',
            'endpoint': '/robot/disconnect',
            'method': 'POST',
            'input': '{\n  "robotId": "arm1"\n}\nHeaders:\n{\n  "Authorization": "Bearer internal_token"\n}',
            'expected_output': '{\n  "status": "ok"\n}',
            'status_code': '200, 400, 401, 500',
            'skip': 'Yes',
            'skip_reason': '需要真实的内部API token和机器人ID',
            'priority': 'Normal'
        },
        {
            'id': 'TO-013',
            'service': 'Tele-Op',
            'name': 'Robot Free',
            'endpoint': '/robot/free',
            'method': 'POST',
            'input': '{\n  "robotId": "arm1"\n}\nHeaders:\n{\n  "Authorization": "Bearer internal_token"\n}',
            'expected_output': '{\n  "status": "ok"\n}',
            'status_code': '200, 400, 401, 500',
            'skip': 'Yes',
            'skip_reason': '需要真实的内部API token和机器人ID',
            'priority': 'Normal'
        },
        {
            'id': 'TO-014',
            'service': 'Tele-Op',
            'name': 'Vision Dolls Compare',
            'endpoint': '/vision/dolls_compare',
            'method': 'POST',
            'input': '{\n  "robotId": "arm1",\n  "controlToken": "token",\n  "views": {\n    "cam1": {\n      "start": ["base64_image"],\n      "end": ["base64_image"]\n    }\n  }\n}\nHeaders:\n{\n  "Authorization": "Bearer internal_token"\n}',
            'expected_output': '{\n  "success": true,\n  "result": {...}\n}',
            'status_code': '200, 400, 401, 500',
            'skip': 'Yes',
            'skip_reason': '需要真实的图片数据',
            'priority': 'Normal'
        },
        {
            'id': 'TO-015',
            'service': 'Tele-Op',
            'name': 'Full Queue Flow (Integration)',
            'endpoint': 'Multiple endpoints',
            'method': 'Mixed',
            'input': 'Integration test:\n1. GET /robots/status\n2. POST /queue/join\n3. POST /queue/status\n4. POST /queue/leave',
            'expected_output': 'All steps return success status codes',
            'status_code': '200 for all steps',
            'skip': 'Yes',
            'skip_reason': '需要真实环境和完整的用户数据',
            'priority': 'Critical'
        },
    ]
    
    # 填充数据
    for row_num, test_case in enumerate(test_cases, 2):
        ws.cell(row=row_num, column=1, value=test_case['id'])
        ws.cell(row=row_num, column=2, value=test_case['service'])
        ws.cell(row=row_num, column=3, value=test_case['name'])
        ws.cell(row=row_num, column=4, value=test_case['endpoint'])
        ws.cell(row=row_num, column=5, value=test_case['method'])
        ws.cell(row=row_num, column=6, value=test_case['input'])
        ws.cell(row=row_num, column=7, value=test_case['expected_output'])
        ws.cell(row=row_num, column=8, value=test_case['status_code'])
        ws.cell(row=row_num, column=9, value=test_case['skip'])
        ws.cell(row=row_num, column=10, value=test_case['skip_reason'])
        ws.cell(row=row_num, column=11, value=test_case['priority'])
        
        # 应用样式
        for col_num in range(1, 12):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = cell_alignment
            cell.border = border
            
            # Skip状态的单元格用黄色标记
            if col_num == 9 and test_case['skip'] == 'Yes':
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # 调整列宽
    column_widths = {
        'A': 10,  # ID
        'B': 20,  # Service
        'C': 30,  # Test Case Name
        'D': 35,  # API Endpoint
        'E': 10,  # Method
        'F': 45,  # Input Parameters
        'G': 45,  # Expected Output
        'H': 20,  # Status Code
        'I': 8,   # Skip
        'J': 35,  # Skip Reason
        'K': 12,  # Priority
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 30
    for row in range(2, len(test_cases) + 2):
        ws.row_dimensions[row].height = 80
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    filename = f'Test_Cases_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    print(f"✅ Excel文件已生成: {filename}")
    print(f"📊 总测试用例数: {len(test_cases)}")
    
    # 统计信息
    skip_count = sum(1 for tc in test_cases if tc['skip'] == 'Yes')
    no_skip_count = len(test_cases) - skip_count
    print(f"✅ 可执行用例: {no_skip_count}")
    print(f"⏭️  跳过用例: {skip_count}")
    
    return filename

if __name__ == '__main__':
    create_test_cases_excel()

